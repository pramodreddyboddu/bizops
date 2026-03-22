"""Product catalog extractor — pull product info from vendor emails and files.

Two modes:
1. Email extraction: Scan past Gmail invoices/orders to find product names,
   quantities, and unit prices from vendor emails.
2. File import: Load products from CSV or Excel files.
"""

from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

from bizops.utils.config import BizOpsConfig, ProductItem


class ProductExtractor:
    """Extract product catalog data from emails and files."""

    def __init__(self, config: BizOpsConfig):
        self.config = config

    # ──────────────────────────────────────────────────────────
    #  Email-based extraction
    # ──────────────────────────────────────────────────────────

    def extract_from_emails(
        self,
        emails: list[dict[str, Any]],
        vendor_name: str | None = None,
    ) -> list[dict[str, Any]]:
        """Extract product line items from vendor emails.

        Parses email bodies for tabular product data — item names,
        quantities, unit prices, and totals.

        Args:
            emails: List of email dicts (from storage or Gmail).
            vendor_name: If set, only process emails from this vendor.

        Returns:
            List of extracted product dicts with name, unit, unit_cost,
            quantity, vendor, source_date.
        """
        products: list[dict[str, Any]] = []
        seen_keys: set[str] = set()

        for email in emails:
            # Filter by vendor if specified
            if vendor_name:
                email_vendor = (email.get("vendor") or "").lower()
                if vendor_name.lower() not in email_vendor:
                    # Also check sender
                    sender = (email.get("sender") or email.get("source_email") or "").lower()
                    if not self._vendor_matches_sender(vendor_name, sender):
                        continue

            body = email.get("body") or ""
            subject = email.get("subject") or ""
            date = email.get("date") or ""
            v_name = email.get("vendor") or "Unknown"

            # Try to extract line items from body
            items = self._extract_line_items(body)

            if not items:
                # Try from subject (single-item orders)
                items = self._extract_from_subject(subject, body)

            for item in items:
                # Dedup by product name + vendor
                key = f"{item['name'].lower()}|{v_name.lower()}"
                if key not in seen_keys:
                    seen_keys.add(key)
                    item["vendor"] = v_name
                    item["source_date"] = date
                    products.append(item)
                elif item.get("unit_cost", 0) > 0:
                    # Update price if we have a newer one
                    for existing in products:
                        if f"{existing['name'].lower()}|{existing['vendor'].lower()}" == key:
                            if date > existing.get("source_date", ""):
                                existing["unit_cost"] = item["unit_cost"]
                                existing["source_date"] = date

        return products

    def _extract_line_items(self, body: str) -> list[dict[str, Any]]:
        """Extract product line items from email body text.

        Looks for patterns like:
        - "Cilantro  50 bunch  $0.75  $37.50"
        - "Cilantro x 50 @ $0.75"
        - "50 lb Onions $0.50/lb"
        - "Item: Cilantro, Qty: 50, Price: $0.75"
        """
        items = []
        if not body:
            return items

        lines = body.split("\n")

        for line in lines:
            line = line.strip()
            if not line or len(line) < 5:
                continue

            # Skip headers and summary lines
            if self._is_header_or_summary(line):
                continue

            # Pattern 1: "ProductName  Qty  Unit  $Price  $Total"
            # e.g., "Cilantro    50  bunch  $0.75  $37.50"
            m = re.match(
                r'^([A-Za-z][A-Za-z\s/&\'-]{2,30}?)\s+'
                r'([\d,.]+)\s+'
                r'(case|lb|lbs|bag|bags|bunch|bunches|each|ea|box|boxes|ct|pk|pack|pcs|gal|oz)\s+'
                r'\$?([\d,.]+)',
                line, re.IGNORECASE,
            )
            if m:
                items.append({
                    "name": m.group(1).strip(),
                    "quantity": self._parse_number(m.group(2)),
                    "unit": self._normalize_unit(m.group(3)),
                    "unit_cost": self._parse_number(m.group(4)),
                })
                continue

            # Pattern 2: "Qty x ProductName @ $Price"
            # e.g., "50 x Cilantro @ $0.75"
            m = re.match(
                r'^([\d,.]+)\s*x\s+'
                r'([A-Za-z][A-Za-z\s/&\'-]{2,30}?)\s+'
                r'[@]\s*\$?([\d,.]+)',
                line, re.IGNORECASE,
            )
            if m:
                items.append({
                    "name": m.group(2).strip(),
                    "quantity": self._parse_number(m.group(1)),
                    "unit": "each",
                    "unit_cost": self._parse_number(m.group(3)),
                })
                continue

            # Pattern 3: "Qty Unit ProductName $Price/unit"
            # e.g., "50 lb Onions $0.50/lb"
            m = re.match(
                r'^([\d,.]+)\s+'
                r'(case|lb|lbs|bag|bags|bunch|bunches|each|ea|box|boxes|ct|pk|pack|gal|oz)\s+'
                r'([A-Za-z][A-Za-z\s/&\'-]{2,30}?)\s+'
                r'\$?([\d,.]+)',
                line, re.IGNORECASE,
            )
            if m:
                items.append({
                    "name": m.group(3).strip(),
                    "quantity": self._parse_number(m.group(1)),
                    "unit": self._normalize_unit(m.group(2)),
                    "unit_cost": self._parse_number(m.group(4)),
                })
                continue

            # Pattern 4: "Item: X, Qty: N, Price: $P" (structured format)
            m = re.search(
                r'(?:item|product|name)\s*[:=]\s*([^,]+?)\s*'
                r'(?:,\s*)?(?:qty|quantity)\s*[:=]\s*([\d,.]+)\s*'
                r'(?:,\s*)?(?:price|cost|rate)\s*[:=]\s*\$?([\d,.]+)',
                line, re.IGNORECASE,
            )
            if m:
                items.append({
                    "name": m.group(1).strip(),
                    "quantity": self._parse_number(m.group(2)),
                    "unit": "each",
                    "unit_cost": self._parse_number(m.group(3)),
                })
                continue

            # Pattern 5: "ProductName  $Price" (simple price list, no qty)
            m = re.match(
                r'^([A-Za-z][A-Za-z\s/&\'-]{2,30}?)\s+'
                r'\$\s*([\d,.]+)\s*'
                r'(?:/?(?:lb|case|each|bag|bunch|box|ct|pk|ea|gal|oz))?',
                line, re.IGNORECASE,
            )
            if m:
                name = m.group(1).strip()
                # Avoid matching generic text
                if not self._is_generic_text(name):
                    unit_match = re.search(
                        r'/(lb|case|each|bag|bunch|box|ct|pk|ea|gal|oz)',
                        line, re.IGNORECASE,
                    )
                    items.append({
                        "name": name,
                        "quantity": 0,
                        "unit": self._normalize_unit(unit_match.group(1)) if unit_match else "each",
                        "unit_cost": self._parse_number(m.group(2)),
                    })

        return items

    def _extract_from_subject(self, subject: str, body: str) -> list[dict[str, Any]]:
        """Extract product info from email subject + body for simple orders."""
        items = []

        # Look for Zelle-style: "Payment for Cilantro order"
        m = re.search(r'(?:payment|order)\s+(?:for\s+)?(.+?)(?:\s+order)?$', subject, re.IGNORECASE)
        if m:
            product_hint = m.group(1).strip()
            # Try to find price in body
            prices = re.findall(r'\$\s*([\d,]+\.?\d{0,2})', body)
            if prices and not self._is_generic_text(product_hint):
                items.append({
                    "name": product_hint,
                    "quantity": 0,
                    "unit": "each",
                    "unit_cost": 0,  # Don't guess from total
                })

        return items

    # ──────────────────────────────────────────────────────────
    #  File-based import
    # ──────────────────────────────────────────────────────────

    def import_from_file(self, file_path: Path) -> list[dict[str, Any]]:
        """Import products from a CSV or Excel file.

        Expected columns (flexible naming):
        - name/product/item (required)
        - unit/uom (optional, default "each")
        - cost/price/unit_cost (optional, default 0)
        - par/par_level/minimum (optional, default 0)
        - multiple/order_multiple (optional, default 1)
        - category (optional, default "food_supplies")
        - sku/item_code (optional)
        - vendor (optional — used when importing for multiple vendors)

        Returns:
            List of product dicts.
        """
        suffix = file_path.suffix.lower()
        if suffix == ".csv":
            return self._import_csv(file_path)
        elif suffix in (".xlsx", ".xls"):
            return self._import_excel(file_path)
        else:
            raise ValueError(f"Unsupported file format: {suffix}. Use .csv or .xlsx")

    def _import_csv(self, path: Path) -> list[dict[str, Any]]:
        """Import products from CSV file."""
        products = []
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            if reader.fieldnames is None:
                raise ValueError("CSV file has no headers.")

            # Build column mapping
            col_map = self._map_columns(reader.fieldnames)
            if "name" not in col_map:
                raise ValueError(
                    "CSV must have a 'name', 'product', or 'item' column. "
                    f"Found: {', '.join(reader.fieldnames)}"
                )

            for row in reader:
                product = self._row_to_product(row, col_map)
                if product:
                    products.append(product)

        return products

    def _import_excel(self, path: Path) -> list[dict[str, Any]]:
        """Import products from Excel file."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl is required for Excel import. Install with: pip install openpyxl")

        wb = load_workbook(path, read_only=True)
        ws = wb.active
        if ws is None:
            raise ValueError("Excel file has no active sheet.")

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        # First row is headers
        headers = [str(h).strip() if h else "" for h in rows[0]]
        col_map = self._map_columns(headers)

        if "name" not in col_map:
            raise ValueError(
                "Excel must have a 'name', 'product', or 'item' column. "
                f"Found: {', '.join(headers)}"
            )

        products = []
        for row_vals in rows[1:]:
            row = dict(zip(headers, row_vals))
            product = self._row_to_product(row, col_map)
            if product:
                products.append(product)

        wb.close()
        return products

    # ──────────────────────────────────────────────────────────
    #  Convert to ProductItem
    # ──────────────────────────────────────────────────────────

    def to_product_items(
        self,
        extracted: list[dict[str, Any]],
        default_category: str = "food_supplies",
    ) -> list[ProductItem]:
        """Convert extracted product dicts to ProductItem models.

        Deduplicates by name (case-insensitive), keeping latest price.
        """
        seen: dict[str, ProductItem] = {}

        for item in extracted:
            name = item.get("name", "").strip()
            if not name:
                continue

            key = name.lower()
            product = ProductItem(
                name=name,
                sku=item.get("sku", ""),
                unit=self._normalize_unit(item.get("unit", "each")),
                unit_cost=float(item.get("unit_cost", 0) or 0),
                par_level=float(item.get("par_level", 0) or 0),
                order_multiple=float(item.get("order_multiple", 1) or 1),
                category=item.get("category", default_category),
            )

            if key not in seen or product.unit_cost > 0:
                seen[key] = product

        return list(seen.values())

    # ──────────────────────────────────────────────────────────
    #  Internal helpers
    # ──────────────────────────────────────────────────────────

    def _vendor_matches_sender(self, vendor_name: str, sender: str) -> bool:
        """Check if a vendor name matches an email sender."""
        for vc in self.config.vendors:
            if vc.name.lower() == vendor_name.lower():
                return vc.matches_email(sender)
        return False

    def _map_columns(self, headers: list[str]) -> dict[str, str]:
        """Map flexible column names to standard field names."""
        mapping: dict[str, str] = {}
        header_lower = {h: h.lower().strip() for h in headers}

        # Name column (required)
        for h, low in header_lower.items():
            if low in ("name", "product", "product name", "item", "item name", "description"):
                mapping["name"] = h
                break

        # Unit column
        for h, low in header_lower.items():
            if low in ("unit", "uom", "unit of measure", "measure"):
                mapping["unit"] = h
                break

        # Cost column
        for h, low in header_lower.items():
            if low in ("cost", "price", "unit cost", "unit price", "rate", "unit_cost"):
                mapping["cost"] = h
                break

        # Par level column
        for h, low in header_lower.items():
            if low in ("par", "par level", "par_level", "minimum", "min", "min_qty"):
                mapping["par"] = h
                break

        # Order multiple
        for h, low in header_lower.items():
            if low in ("multiple", "order multiple", "order_multiple", "case size", "case_size"):
                mapping["multiple"] = h
                break

        # Category
        for h, low in header_lower.items():
            if low in ("category", "cat", "expense category", "type"):
                mapping["category"] = h
                break

        # SKU
        for h, low in header_lower.items():
            if low in ("sku", "item code", "item_code", "code", "product code"):
                mapping["sku"] = h
                break

        # Vendor
        for h, low in header_lower.items():
            if low in ("vendor", "supplier", "vendor name"):
                mapping["vendor"] = h
                break

        return mapping

    def _row_to_product(
        self,
        row: dict[str, Any],
        col_map: dict[str, str],
    ) -> dict[str, Any] | None:
        """Convert a row dict to a product dict using column mapping."""
        name_col = col_map.get("name", "")
        name = str(row.get(name_col, "")).strip() if name_col else ""

        if not name:
            return None

        return {
            "name": name,
            "sku": str(row.get(col_map.get("sku", ""), "") or "").strip(),
            "unit": str(row.get(col_map.get("unit", ""), "each") or "each").strip(),
            "unit_cost": self._parse_cost(row.get(col_map.get("cost", ""), 0)),
            "par_level": self._parse_number(row.get(col_map.get("par", ""), 0)),
            "order_multiple": self._parse_number(row.get(col_map.get("multiple", ""), 1)) or 1,
            "category": str(row.get(col_map.get("category", ""), "food_supplies") or "food_supplies").strip(),
            "vendor": str(row.get(col_map.get("vendor", ""), "") or "").strip(),
        }

    @staticmethod
    def _parse_number(val: Any) -> float:
        """Parse a number from various formats."""
        if isinstance(val, (int, float)):
            return float(val)
        if not val:
            return 0.0
        try:
            return float(str(val).replace(",", "").replace("$", "").strip())
        except (ValueError, TypeError):
            return 0.0

    @staticmethod
    def _parse_cost(val: Any) -> float:
        """Parse a cost/price value, handling $ and comma formatting."""
        if isinstance(val, (int, float)):
            return round(float(val), 2)
        if not val:
            return 0.0
        try:
            cleaned = str(val).replace(",", "").replace("$", "").strip()
            return round(float(cleaned), 2)
        except (ValueError, TypeError):
            return 0.0

    @staticmethod
    def _normalize_unit(unit: str) -> str:
        """Normalize unit names to standard forms."""
        if not unit:
            return "each"
        unit_lower = unit.lower().strip()
        unit_map = {
            "lbs": "lb", "pounds": "lb", "pound": "lb",
            "bags": "bag",
            "bunches": "bunch",
            "boxes": "box",
            "ea": "each",
            "pcs": "each", "pieces": "each", "piece": "each",
            "pk": "pack", "packs": "pack",
            "cs": "case", "cases": "case",
            "gals": "gal", "gallon": "gal", "gallons": "gal",
            "ounce": "oz", "ounces": "oz",
        }
        return unit_map.get(unit_lower, unit_lower)

    @staticmethod
    def _is_header_or_summary(line: str) -> bool:
        """Check if a line is a table header or summary rather than a product."""
        lower = line.lower().strip()
        # If the line has "key: value" structured data, it's not a header
        if re.search(r'(?:item|product|name)\s*[:=]\s*.+(?:qty|quantity)\s*[:=]', lower):
            return False

        skip_patterns = [
            "subtotal", "sub total", "total", "tax", "shipping",
            "discount", "grand total", "balance", "amount due",
            "item", "product", "description", "qty", "quantity",
            "price", "unit cost", "---", "===", "***",
        ]
        return any(lower.startswith(p) or lower == p for p in skip_patterns)

    @staticmethod
    def _is_generic_text(text: str) -> bool:
        """Check if text is too generic to be a product name."""
        generic = {
            "total", "subtotal", "tax", "shipping", "discount",
            "payment", "invoice", "order", "receipt", "thank you",
            "thank", "please", "note", "notes", "description",
            "amount", "balance", "due", "from", "to", "date",
        }
        return text.lower().strip() in generic
