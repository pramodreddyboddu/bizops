"""Excel export helpers for invoices and expenses."""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────
#  Transaction type classification
# ──────────────────────────────────────────────────────────────

def classify_transaction(inv: dict[str, Any]) -> str:
    """Classify an invoice into: payment, deposit, order, or other.

    Classification from Desi Delight's perspective:
      payment = money LEAVING Desi Delight (you paid someone)
      deposit = money COMING IN to Desi Delight (someone paid you)
      order   = purchase orders, statements, confirmations (no cash movement)
      other   = spam, promos, uncategorized

    Returns one of: 'payment', 'deposit', 'order', 'other'.
    """
    subject = (inv.get("subject") or "").lower()

    # ── SPAM / PROMO / NOISE FILTER (catch first) ──
    # Samsung/carrier promos with fake amounts
    if "pre-order" in subject and ("savings" in subject or "ghost" in subject):
        return "other"
    # Alibaba spam
    if "order your samples" in subject or "order with ease" in subject:
        return "other"
    # Google Play personal purchases
    if "google play" in subject and "order receipt" in subject:
        return "other"
    # Bank statement notifications — no data, just "go log in"
    if "your statement is available" in subject:
        return "other"

    # ── PAYMENTS OUT (Money LEAVING Desi Delight) ──
    # Zelle outbound payments
    if "zelle" in subject and "has been sent" in subject:
        return "payment"
    # Scheduled payments to vendors (Bank of America)
    if "payment" in subject and "has been scheduled" in subject:
        return "payment"
    # Om Produce RECEIVED payment FROM Desi Delight = you PAID them = money OUT
    if "has received a payment" in subject and "from desi delight" in subject:
        return "payment"
    # Om Produce Payment Initiated = money OUT
    if "payment initiated" in subject:
        return "payment"
    # Recurring payment reminders (upcoming money out)
    if "recurring payment" in subject:
        return "payment"
    # Invoice due (money you owe)
    if "invoice" in subject and "due" in subject:
        return "payment"
    # Invoice from vendor (e.g. "Invoice 451605 from Superior Trading")
    if re.search(r"invoice\s+\d+\s+from", subject):
        return "payment"

    # ── DEPOSITS IN (Money COMING INTO Desi Delight's bank) ──
    # "Your DoorDash payment for..." = ACTUAL bank deposit (net after commission)
    if "your doordash payment" in subject:
        return "deposit"

    # ── DOORDASH SALES SUMMARY (not actual deposit — gross sales before commission) ──
    # "DoorDash Payment to Desi Delight..." = gross sales total, NOT bank deposit
    if "doordash payment to" in subject and "desi delight" in subject:
        return "order"  # Treat as informational / sales summary

    # DoorDash financial statement = informational
    if "financial statement" in subject and "merchant" in subject:
        return "order"

    # ── UTILITY BILLS (Money OUT — recurring business expenses) ──
    if "gexa" in subject and ("invoice" in subject or "e-invoice" in subject):
        return "payment"
    if "atmos" in subject and "bill" in subject:
        return "payment"
    if "at&t" in subject and "bill" in subject:
        return "payment"
    princeton_keywords = ("bill", "utility", "water", "city")
    if "princeton" in subject and any(kw in subject for kw in princeton_keywords):
        return "payment"

    # ── ORDERS (Purchase orders, statements, confirmations — no cash movement) ──
    if "order form" in subject:
        return "order"
    if "sales order" in subject:
        return "order"
    if re.search(r"order\s*\(#\w+\)\s*received", subject):
        return "order"
    if "delivery confirmation" in subject:
        return "order"
    if "a/r statement" in subject:
        return "order"
    if "reconciliation statement" in subject:
        return "order"
    # Bills from non-utility sources (not caught above)
    if "bill" in subject and ("ready" in subject or "available" in subject):
        # Check if it's a utility we missed — if sender has utility keywords
        return "payment"
    if "e-invoice" in subject:
        return "order"
    if "order receipt" in subject:
        return "order"
    # General statements (bank, vendor)
    if "statement" in subject and "zelle" not in subject:
        return "order"
    # Order links / web store
    if "order link" in subject or "web store" in subject:
        return "order"

    return "other"


def extract_zelle_recipient(subject: str) -> str:
    """Extract the Zelle recipient name from a Zelle payment subject line.

    Example: 'Zelle® payment of $1,248.31 to YAMAN HALAL MEAT LLC has been sent'
             → 'YAMAN HALAL MEAT LLC'
    """
    match = re.search(r"to\s+(.+?)\s+has been sent", subject, re.IGNORECASE)
    if match:
        return match.group(1).strip()
    return ""


def deduplicate_invoices(invoices: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Remove duplicate emails based on subject + date + amount.

    Keeps the first occurrence of each unique (subject, date, amount) combo.
    """
    seen: set[str] = set()
    unique: list[dict[str, Any]] = []

    for inv in invoices:
        # Build a dedup key from subject + date + amount
        key = f"{inv.get('subject', '')}|{inv.get('date', '')}|{inv.get('amount', 0)}"
        if key not in seen:
            seen.add(key)
            unique.append(inv)

    return unique


def segregate_invoices(
    invoices: list[dict[str, Any]],
) -> dict[str, list[dict[str, Any]]]:
    """Deduplicate, classify, and split invoices into payments, deposits, orders.

    Spam/promos (classified as 'other') are dropped entirely.

    From Desi Delight's perspective:
      payment = money OUT (Zelle sent, scheduled payments, Om Produce received FROM you)
      deposit = money IN  (DoorDash payouts TO you)
      order   = no cash movement (order forms, statements, confirmations)
    """
    # Step 1: Remove duplicates
    unique_invoices = deduplicate_invoices(invoices)

    buckets: dict[str, list[dict[str, Any]]] = {
        "payment": [],
        "deposit": [],
        "order": [],
    }

    for inv in unique_invoices:
        tx_type = classify_transaction(inv)
        enriched = {**inv, "transaction_type": tx_type}
        subject = (inv.get("subject") or "").lower()

        # ── Enrich vendor names for payments ──
        # Zelle: extract recipient name
        if tx_type == "payment" and "zelle" in subject and "has been sent" in subject:
            recipient = extract_zelle_recipient(inv.get("subject", ""))
            if recipient:
                enriched["vendor"] = recipient
                enriched["category"] = "zelle_payment"

        # Scheduled: extract vendor name
        if tx_type == "payment" and "has been scheduled" in subject:
            match = re.search(r"payment to\s+(.+?)\s+has been scheduled",
                              inv.get("subject", ""), re.IGNORECASE)
            if match:
                enriched["vendor"] = match.group(1).strip()
                enriched["category"] = "scheduled_payment"

        # Om Produce received payment from you = payment OUT
        if tx_type == "payment" and "has received a payment" in subject:
            enriched["vendor"] = "Om Produce"
            enriched["category"] = "vendor_payment"

        # Om Produce Payment Initiated
        if tx_type == "payment" and "payment initiated" in subject:
            enriched["vendor"] = "Om Produce"
            enriched["category"] = "vendor_payment"

        # Invoice from vendor
        if tx_type == "payment":
            inv_match = re.search(r"invoice\s+\d+\s+from\s+(.+)", subject)
            if inv_match:
                enriched["vendor"] = inv_match.group(1).strip().title()
                enriched["category"] = "vendor_invoice"

        # ── Enrich deposit sources ──
        if tx_type == "deposit" and "doordash" in subject:
            enriched["vendor"] = "DoorDash"
            enriched["category"] = "doordash_payout"
            enriched["status"] = "deposited"

        # ── Enrich DoorDash sales summaries (in orders) ──
        if tx_type == "order" and "doordash payment to" in subject:
            enriched["vendor"] = "DoorDash"
            enriched["category"] = "doordash_sales_summary"

        if tx_type == "order" and "financial statement" in subject and "merchant" in subject:
            enriched["vendor"] = "DoorDash"
            enriched["category"] = "doordash_statement"

        # ── Enrich utility bills ──
        if "gexa" in subject:
            enriched["vendor"] = "Gexa Energy"
            enriched["category"] = "utilities"
        if "atmos" in subject:
            enriched["vendor"] = "Atmos Energy"
            enriched["category"] = "utilities"
        if "at&t" in subject and "bill" in subject:
            enriched["vendor"] = "AT&T"
            enriched["category"] = "utilities"

        # ── Enrich vendor statements ──
        if "rrk foods" in subject:
            enriched["vendor"] = "RRK Foods"
            enriched["category"] = "grocery_supplies"
        if "desi delight market place" in subject and "statement" in subject:
            enriched["vendor"] = "House of Spices"
            enriched["category"] = "grocery_supplies"
        if "zeenat" in subject:
            enriched["vendor"] = "Zeenat Inc"
            enriched["category"] = "grocery_supplies"
        if "united trading" in subject:
            enriched["vendor"] = "United Trading"
            enriched["category"] = "grocery_supplies"

        # ── Enrich Straunt customer orders ──
        source = (inv.get("source_email") or "").lower()
        if "straunt" in source:
            enriched["vendor"] = "Straunt (Customer Orders)"
            enriched["category"] = "customer_orders"

        # ── Jana Food invoice forwarded by staff ──
        if "inv-jf" in subject or "jana food" in subject:
            enriched["vendor"] = "Jana Food Services"
            enriched["category"] = "food_supplies"

        # Skip spam/promos — don't include in any report
        if tx_type == "other":
            continue

        buckets[tx_type].append(enriched)

    return buckets


# ──────────────────────────────────────────────────────────────
#  Styling helpers
# ──────────────────────────────────────────────────────────────

# Sheet color themes
SHEET_THEMES = {
    "payment":  {"fill": "8B0000", "name": "Payments OUT",  "tab": "FF4444"},  # Red
    "deposit":  {"fill": "1B5E20", "name": "Deposits IN",   "tab": "4CAF50"},  # Green
    "order":    {"fill": "0D47A1", "name": "Orders",        "tab": "2196F3"},  # Blue
    "other":    {"fill": "4E342E", "name": "Other",         "tab": "795548"},  # Brown
    "summary":  {"fill": "2F5233", "name": "Summary",       "tab": "2F5233"},  # DD Green
}


def _get_styles(theme_color: str):
    """Return consistent styles for a given header color."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=theme_color, end_color=theme_color, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    return header_font, header_fill, header_alignment, thin_border


# ──────────────────────────────────────────────────────────────
#  Sheet builders
# ──────────────────────────────────────────────────────────────

def _write_sheet(
    ws,
    items: list[dict[str, Any]],
    title: str,
    theme_color: str,
    headers: list[str],
    row_builder,
    start_date: str,
    end_date: str,
):
    """Write a formatted sheet with header, data rows, and totals."""
    header_font, header_fill, header_alignment, thin_border = _get_styles(theme_color)
    num_cols = len(headers)

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1, value=f"{title}  ({start_date} to {end_date})")
    title_cell.font = Font(bold=True, size=14, color=theme_color)
    title_cell.alignment = Alignment(horizontal="center")

    # Count row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    count_cell = ws.cell(row=2, column=1, value=f"{len(items)} transactions")
    count_cell.font = Font(italic=True, size=10, color="666666")
    count_cell.alignment = Alignment(horizontal="center")

    # Headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row_idx, inv in enumerate(items, 5):
        values = row_builder(inv)
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            # Format amount columns as currency
            if headers[col - 1] == "Amount":
                cell.number_format = '$#,##0.00'
            # Color status
            if headers[col - 1] == "Status":
                if str(val).lower() == "unpaid":
                    cell.font = Font(color="CC0000")
                elif str(val).lower() == "paid":
                    cell.font = Font(color="006600")

    # Total row
    if items:
        last_row = len(items) + 5
        amt_col = None
        for i, h in enumerate(headers):
            if h == "Amount":
                amt_col = i + 1
                break

        if amt_col:
            ws.cell(row=last_row, column=amt_col - 1, value="TOTAL").font = Font(bold=True, size=12)
            total_cell = ws.cell(
                row=last_row, column=amt_col,
                value=f"=SUM({get_column_letter(amt_col)}5:{get_column_letter(amt_col)}{last_row - 1})"
            )
            total_cell.font = Font(bold=True, size=12)
            total_cell.number_format = '$#,##0.00'
            total_cell.border = thin_border

    # Auto-width
    for col in range(1, num_cols + 1):
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(4, len(items) + 6)),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 35)


def _build_summary_sheet(
    ws,
    buckets: dict[str, list[dict[str, Any]]],
    start_date: str,
    end_date: str,
):
    """Build the Summary dashboard sheet."""
    theme = SHEET_THEMES["summary"]
    header_font, header_fill, header_alignment, thin_border = _get_styles(theme["fill"])

    # Title
    ws.merge_cells("A1:E1")
    title = ws.cell(row=1, column=1, value=f"Desi Delight — Financial Summary ({start_date} to {end_date})")
    title.font = Font(bold=True, size=16, color="2F5233")
    title.alignment = Alignment(horizontal="center")

    # Cash flow summary
    payments_total = sum(inv.get("amount") or 0 for inv in buckets.get("payment", []))
    deposits_total = sum(inv.get("amount") or 0 for inv in buckets.get("deposit", []))
    net = deposits_total - payments_total

    row = 3
    # Payments OUT
    ws.cell(row=row, column=1, value="Payments OUT").font = Font(bold=True, size=12, color="CC0000")
    ws.cell(row=row, column=2, value=f"{len(buckets.get('payment', []))} txns").font = Font(italic=True, color="666666")
    pay_cell = ws.cell(row=row, column=3, value=payments_total)
    pay_cell.number_format = '$#,##0.00'
    pay_cell.font = Font(bold=True, size=13, color="CC0000")
    for c in range(1, 4):
        ws.cell(row=row, column=c).border = thin_border

    row += 1
    # Deposits IN
    ws.cell(row=row, column=1, value="Deposits IN").font = Font(bold=True, size=12, color="006600")
    ws.cell(row=row, column=2, value=f"{len(buckets.get('deposit', []))} txns").font = Font(italic=True, color="666666")
    dep_cell = ws.cell(row=row, column=3, value=deposits_total)
    dep_cell.number_format = '$#,##0.00'
    dep_cell.font = Font(bold=True, size=13, color="006600")
    for c in range(1, 4):
        ws.cell(row=row, column=c).border = thin_border

    row += 2
    # Net cash flow
    ws.cell(row=row, column=1, value="NET CASH FLOW").font = Font(bold=True, size=14)
    ws.cell(row=row, column=2, value="(Deposits - Payments)").font = Font(italic=True, color="666666")
    net_cell = ws.cell(row=row, column=3, value=net)
    net_cell.number_format = '$#,##0.00'
    net_color = "006600" if net >= 0 else "CC0000"
    net_cell.font = Font(bold=True, size=14, color=net_color)

    # Top vendors by payment
    row += 3
    ws.cell(row=row, column=1, value="Top Vendor Payments").font = Font(bold=True, size=12, color="2F5233")
    row += 1
    vendor_totals: dict[str, float] = {}
    for inv in buckets.get("payment", []):
        v = inv.get("vendor", "Unknown")
        vendor_totals[v] = vendor_totals.get(v, 0) + (inv.get("amount") or 0)

    for v, total in sorted(vendor_totals.items(), key=lambda x: -x[1]):
        ws.cell(row=row, column=1, value=v)
        amt = ws.cell(row=row, column=3, value=total)
        amt.number_format = '$#,##0.00'
        row += 1

    # Auto-width
    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 25


# ──────────────────────────────────────────────────────────────
#  Main export function
# ──────────────────────────────────────────────────────────────

def export_invoices_to_excel(
    invoices: list[dict[str, Any]],
    config: Any,
    start_date: str,
    end_date: str,
    output_path: Path | None = None,
) -> Path:
    """Export invoices to a multi-sheet Excel workbook, segregated by type."""
    if output_path is None:
        month_str = datetime.now().strftime("%Y_%m")
        filename = f"DD_INVOICES_{month_str}.xlsx"
        output_path = config.output_dir / filename

    config.ensure_dirs()

    # Segregate
    buckets = segregate_invoices(invoices)

    wb = Workbook()

    # ── Sheet 1: Summary Dashboard ──
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.sheet_properties.tabColor = SHEET_THEMES["summary"]["tab"]
    _build_summary_sheet(ws_summary, buckets, start_date, end_date)

    # ── Sheet 2: Payments OUT ──
    if buckets["payment"]:
        ws_pay = wb.create_sheet("Payments OUT")
        ws_pay.sheet_properties.tabColor = SHEET_THEMES["payment"]["tab"]
        _write_sheet(
            ws_pay,
            sorted(buckets["payment"], key=lambda x: x.get("date", ""), reverse=True),
            "💸 Payments OUT (Zelle & Scheduled)",
            SHEET_THEMES["payment"]["fill"],
            ["#", "Date", "Vendor / Recipient", "Amount", "Status", "Category", "Subject"],
            lambda inv: [
                "",
                inv.get("date", ""),
                inv.get("vendor", "Unknown"),
                inv.get("amount") or 0,
                (inv.get("status") or "unknown").capitalize(),
                inv.get("category", ""),
                (inv.get("subject") or "")[:60],
            ],
            start_date,
            end_date,
        )
        # Add row numbers
        for i, row in enumerate(range(5, 5 + len(buckets["payment"]))):
            ws_pay.cell(row=row, column=1, value=i + 1)

    # ── Sheet 3: Deposits IN ──
    if buckets["deposit"]:
        ws_dep = wb.create_sheet("Deposits IN")
        ws_dep.sheet_properties.tabColor = SHEET_THEMES["deposit"]["tab"]
        _write_sheet(
            ws_dep,
            sorted(buckets["deposit"], key=lambda x: x.get("date", ""), reverse=True),
            "💰 Deposits IN (DoorDash & Vendor Confirmations)",
            SHEET_THEMES["deposit"]["fill"],
            ["#", "Date", "Source", "Amount", "Status", "Category", "Subject"],
            lambda inv: [
                "",
                inv.get("date", ""),
                inv.get("vendor", "Unknown"),
                inv.get("amount") or 0,
                (inv.get("status") or "unknown").capitalize(),
                inv.get("category", ""),
                (inv.get("subject") or "")[:60],
            ],
            start_date,
            end_date,
        )
        for i, row in enumerate(range(5, 5 + len(buckets["deposit"]))):
            ws_dep.cell(row=row, column=1, value=i + 1)

    # ── Sheet 4: Orders ──
    if buckets["order"]:
        ws_ord = wb.create_sheet("Orders")
        ws_ord.sheet_properties.tabColor = SHEET_THEMES["order"]["tab"]
        _write_sheet(
            ws_ord,
            sorted(buckets["order"], key=lambda x: x.get("date", ""), reverse=True),
            "📦 Orders, Statements & Confirmations",
            SHEET_THEMES["order"]["fill"],
            ["#", "Date", "Vendor", "Amount", "Category", "Subject"],
            lambda inv: [
                "",
                inv.get("date", ""),
                inv.get("vendor", "Unknown"),
                inv.get("amount") or 0,
                inv.get("category", ""),
                (inv.get("subject") or "")[:70],
            ],
            start_date,
            end_date,
        )
        for i, row in enumerate(range(5, 5 + len(buckets["order"]))):
            ws_ord.cell(row=row, column=1, value=i + 1)

    wb.save(output_path)
    return output_path


# ──────────────────────────────────────────────────────────────
#  P&L Workbook Export
# ──────────────────────────────────────────────────────────────

def export_pl_workbook(
    pl_data: dict[str, Any],
    config: Any,
    output_path: Path | None = None,
) -> Path:
    """Export a multi-tab P&L Excel workbook.

    Tabs:
      1. Sales Summary — daily Toast POS data
      2. Expenses — all expenses grouped by category
      3. P&L — revenue vs expenses summary
      4. Vendor Summary — spend per vendor
    """
    if output_path is None:
        period = pl_data.get("period", {})
        start = period.get("start", datetime.now().strftime("%Y-%m-%d"))
        month_str = start[:7].replace("-", "_")
        filename = f"DD_PL_REPORT_{month_str}.xlsx"
        output_path = config.output_dir / filename

    config.ensure_dirs()

    wb = Workbook()
    theme = "2F5233"
    header_font, header_fill, header_alignment, thin_border = _get_styles(theme)

    period = pl_data.get("period", {})
    start_date = period.get("start", "?")
    end_date = period.get("end", "?")

    # ── Tab 1: Sales Summary ──
    ws_sales = wb.active
    ws_sales.title = "Sales Summary"
    ws_sales.sheet_properties.tabColor = "4CAF50"

    ws_sales.merge_cells("A1:G1")
    t = ws_sales.cell(row=1, column=1, value=f"Daily Sales Summary ({start_date} to {end_date})")
    t.font = Font(bold=True, size=14, color=theme)
    t.alignment = Alignment(horizontal="center")

    sales_headers = ["Date", "Gross Sales", "Net Sales", "Tax", "Tips", "Refunds", "Orders"]
    for col, h in enumerate(sales_headers, 1):
        cell = ws_sales.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    daily_sales = pl_data.get("daily_sales", [])
    for row_idx, day in enumerate(sorted(daily_sales, key=lambda x: x.get("date", "")), 4):
        ws_sales.cell(row=row_idx, column=1, value=day.get("date", ""))
        for col, key in enumerate(["gross", "net", "tax", "tips", "refunds"], 2):
            cell = ws_sales.cell(row=row_idx, column=col, value=day.get(key) or 0)
            cell.number_format = '$#,##0.00'
        ws_sales.cell(row=row_idx, column=7, value=day.get("orders") or 0)
        for col in range(1, 8):
            ws_sales.cell(row=row_idx, column=col).border = thin_border

    # Totals row
    if daily_sales:
        last = len(daily_sales) + 4
        ws_sales.cell(row=last, column=1, value="TOTAL").font = Font(bold=True)
        for col in range(2, 7):
            c = ws_sales.cell(row=last, column=col,
                              value=f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{last - 1})")
            c.font = Font(bold=True)
            c.number_format = '$#,##0.00'
        c = ws_sales.cell(row=last, column=7,
                          value=f"=SUM(G4:G{last - 1})")
        c.font = Font(bold=True)

    for col in range(1, 8):
        ws_sales.column_dimensions[get_column_letter(col)].width = 15

    # ── Tab 2: Expenses ──
    ws_exp = wb.create_sheet("Expenses")
    ws_exp.sheet_properties.tabColor = "FF4444"

    ws_exp.merge_cells("A1:F1")
    t = ws_exp.cell(row=1, column=1, value=f"Expenses by Category ({start_date} to {end_date})")
    t.font = Font(bold=True, size=14, color="8B0000")
    t.alignment = Alignment(horizontal="center")

    exp_headers = ["Date", "Vendor", "Description", "Amount", "Category", "Status"]
    for col, h in enumerate(exp_headers, 1):
        cell = ws_exp.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
        cell.alignment = header_alignment
        cell.border = thin_border

    expenses_by_cat = pl_data.get("expenses_by_category", {})
    row = 4
    for cat in sorted(expenses_by_cat.keys()):
        items = expenses_by_cat[cat]
        if not items:
            continue
        # Category header
        ws_exp.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cat_cell = ws_exp.cell(row=row, column=1, value=cat.replace("_", " ").upper())
        cat_cell.font = Font(bold=True, size=11, color="8B0000")
        cat_cell.fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
        row += 1

        for item in sorted(items, key=lambda x: x.get("date", ""), reverse=True):
            ws_exp.cell(row=row, column=1, value=item.get("date", "")).border = thin_border
            ws_exp.cell(row=row, column=2, value=item.get("vendor", "Unknown")).border = thin_border
            ws_exp.cell(row=row, column=3, value=(item.get("subject") or "")[:50]).border = thin_border
            amt = ws_exp.cell(row=row, column=4, value=item.get("amount") or 0)
            amt.number_format = '$#,##0.00'
            amt.border = thin_border
            ws_exp.cell(row=row, column=5, value=cat.replace("_", " ").title()).border = thin_border
            ws_exp.cell(row=row, column=6, value=(item.get("status") or "unknown").capitalize()).border = thin_border
            row += 1

        # Subtotal
        cat_total = sum(i.get("amount") or 0 for i in items)
        ws_exp.cell(row=row, column=3, value=f"Subtotal ({len(items)} items)").font = Font(bold=True, italic=True)
        st = ws_exp.cell(row=row, column=4, value=cat_total)
        st.font = Font(bold=True)
        st.number_format = '$#,##0.00'
        row += 1

    # Grand total
    row += 1
    ws_exp.cell(row=row, column=3, value="GRAND TOTAL").font = Font(bold=True, size=12)
    total_expenses = pl_data.get("totals", {}).get("total_expenses", 0)
    gt = ws_exp.cell(row=row, column=4, value=total_expenses)
    gt.font = Font(bold=True, size=12, color="8B0000")
    gt.number_format = '$#,##0.00'

    for col in range(1, 7):
        ws_exp.column_dimensions[get_column_letter(col)].width = 20

    # ── Tab 3: P&L ──
    ws_pl = wb.create_sheet("P&L")
    ws_pl.sheet_properties.tabColor = "2F5233"

    ws_pl.merge_cells("A1:C1")
    t = ws_pl.cell(row=1, column=1, value=f"Profit & Loss Statement ({start_date} to {end_date})")
    t.font = Font(bold=True, size=16, color=theme)
    t.alignment = Alignment(horizontal="center")

    revenue = pl_data.get("revenue", {})
    row = 3

    # Revenue section
    ws_pl.cell(row=row, column=1, value="REVENUE").font = Font(bold=True, size=13, color="006600")
    row += 1
    rev_items = [
        ("Gross Sales", revenue.get("gross_sales", 0)),
        ("Net Sales", revenue.get("net_sales", 0)),
        ("Tax Collected", revenue.get("tax", 0)),
        ("Tips", revenue.get("tips", 0)),
    ]
    for label, val in rev_items:
        ws_pl.cell(row=row, column=1, value=f"  {label}")
        c = ws_pl.cell(row=row, column=3, value=val or 0)
        c.number_format = '$#,##0.00'
        c.font = Font(color="006600")
        row += 1

    total_revenue = pl_data.get("totals", {}).get("total_revenue", 0)
    ws_pl.cell(row=row, column=1, value="Total Revenue").font = Font(bold=True, size=12)
    tr = ws_pl.cell(row=row, column=3, value=total_revenue)
    tr.font = Font(bold=True, size=12, color="006600")
    tr.number_format = '$#,##0.00'
    row += 2

    # Expenses section
    ws_pl.cell(row=row, column=1, value="EXPENSES").font = Font(bold=True, size=13, color="CC0000")
    row += 1
    for cat in sorted(expenses_by_cat.keys()):
        items = expenses_by_cat[cat]
        cat_total = sum(i.get("amount") or 0 for i in items)
        if cat_total > 0:
            ws_pl.cell(row=row, column=1, value=f"  {cat.replace('_', ' ').title()}")
            ws_pl.cell(row=row, column=2, value=f"({len(items)})")
            c = ws_pl.cell(row=row, column=3, value=cat_total)
            c.number_format = '$#,##0.00'
            c.font = Font(color="CC0000")
            row += 1

    ws_pl.cell(row=row, column=1, value="Total Expenses").font = Font(bold=True, size=12)
    te = ws_pl.cell(row=row, column=3, value=total_expenses)
    te.font = Font(bold=True, size=12, color="CC0000")
    te.number_format = '$#,##0.00'
    row += 2

    # Net Profit
    net_profit = pl_data.get("totals", {}).get("net_profit", 0)
    ws_pl.cell(row=row, column=1, value="NET PROFIT / (LOSS)").font = Font(bold=True, size=14)
    np_color = "006600" if (net_profit or 0) >= 0 else "CC0000"
    np = ws_pl.cell(row=row, column=3, value=net_profit or 0)
    np.font = Font(bold=True, size=14, color=np_color)
    np.number_format = '$#,##0.00'

    for col in range(1, 4):
        ws_pl.column_dimensions[get_column_letter(col)].width = 25

    # ── Tab 4: Vendor Summary ──
    ws_vendor = wb.create_sheet("Vendor Summary")
    ws_vendor.sheet_properties.tabColor = "795548"

    ws_vendor.merge_cells("A1:E1")
    t = ws_vendor.cell(row=1, column=1, value=f"Vendor Spend Summary ({start_date} to {end_date})")
    t.font = Font(bold=True, size=14, color=theme)
    t.alignment = Alignment(horizontal="center")

    v_headers = ["Vendor", "Total Spent", "# Invoices", "Avg Invoice", "Last Invoice"]
    for col, h in enumerate(v_headers, 1):
        cell = ws_vendor.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Aggregate vendor data from all expense categories
    vendor_data: dict[str, dict[str, Any]] = {}
    for cat_items in expenses_by_cat.values():
        for item in cat_items:
            v = item.get("vendor", "Unknown")
            if v not in vendor_data:
                vendor_data[v] = {"total": 0, "count": 0, "last_date": ""}
            vendor_data[v]["total"] += item.get("amount") or 0
            vendor_data[v]["count"] += 1
            d = item.get("date", "")
            if d > vendor_data[v]["last_date"]:
                vendor_data[v]["last_date"] = d

    row = 4
    for v, info in sorted(vendor_data.items(), key=lambda x: -x[1]["total"]):
        ws_vendor.cell(row=row, column=1, value=v).border = thin_border
        tc = ws_vendor.cell(row=row, column=2, value=info["total"])
        tc.number_format = '$#,##0.00'
        tc.border = thin_border
        ws_vendor.cell(row=row, column=3, value=info["count"]).border = thin_border
        avg = info["total"] / info["count"] if info["count"] > 0 else 0
        ac = ws_vendor.cell(row=row, column=4, value=avg)
        ac.number_format = '$#,##0.00'
        ac.border = thin_border
        ws_vendor.cell(row=row, column=5, value=info["last_date"]).border = thin_border
        row += 1

    for col in range(1, 6):
        ws_vendor.column_dimensions[get_column_letter(col)].width = 20

    wb.save(output_path)
    return output_path


# ──────────────────────────────────────────────────────────────
#  Reconciliation workbook export
# ──────────────────────────────────────────────────────────────

def export_reconciliation_workbook(
    result: dict[str, Any],
    cash_flow: dict[str, Any],
    config: Any,
    output_path: Path | None = None,
) -> Path:
    """Export a multi-tab reconciliation Excel workbook.

    Tabs:
      1. Summary — match rate, totals, period info
      2. Matched — side-by-side bank vs invoice
      3. Unmatched Bank — bank transactions with no invoice match
      4. Unmatched Invoices — invoices with no bank transaction
      5. Cash Flow — all bank transactions categorized
    """
    if output_path is None:
        month_str = datetime.now().strftime("%Y_%m")
        filename = f"DD_RECONCILIATION_{month_str}.xlsx"
        output_path = config.output_dir / filename

    config.ensure_dirs()

    wb = Workbook()
    theme = "1565C0"  # Blue theme for bank/reconciliation
    header_font, header_fill, header_alignment, thin_border = _get_styles(theme)

    summary = result.get("summary", {})

    # ── Tab 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "1565C0"

    ws.merge_cells("A1:C1")
    t = ws.cell(row=1, column=1, value="Bank Reconciliation Summary")
    t.font = Font(bold=True, size=16, color=theme)
    t.alignment = Alignment(horizontal="center")

    summary_rows = [
        ("Total Bank Transactions", summary.get("total_bank_txns", 0)),
        ("Total Invoices", summary.get("total_invoices", 0)),
        ("Matched", summary.get("matched_count", 0)),
        ("Match Rate", f"{summary.get('match_rate', 0):.1f}%"),
        ("Unmatched Bank Txns", summary.get("unmatched_bank_count", 0)),
        ("Unmatched Invoices", summary.get("unmatched_invoice_count", 0)),
        ("", ""),
        ("Total Bank Debits", summary.get("total_bank_debits", 0)),
        ("Total Bank Credits", summary.get("total_bank_credits", 0)),
        ("Net Bank Flow", summary.get("net_bank_flow", 0)),
    ]

    for i, (label, val) in enumerate(summary_rows, 3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=i, column=3, value=val)
        if isinstance(val, (int, float)) and label and "Rate" not in label:
            cell.number_format = '$#,##0.00' if isinstance(val, float) else '#,##0'

    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].width = 25

    # ── Tab 2: Matched ──
    ws_matched = wb.create_sheet("Matched")
    ws_matched.sheet_properties.tabColor = "4CAF50"

    ws_matched.merge_cells("A1:H1")
    t = ws_matched.cell(row=1, column=1, value="Matched Transactions")
    t.font = Font(bold=True, size=14, color="2E7D32")
    t.alignment = Alignment(horizontal="center")

    m_headers = ["Bank Date", "Bank Description", "Bank Amount",
                 "Invoice Date", "Invoice Vendor", "Invoice Amount",
                 "Match Type", "Score"]
    for col, h in enumerate(m_headers, 1):
        cell = ws_matched.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, match in enumerate(result.get("matched", []), 4):
        btxn = match.get("bank_txn", {})
        inv = match.get("invoice", {})
        ws_matched.cell(row=row_idx, column=1, value=btxn.get("date", "")).border = thin_border
        ws_matched.cell(row=row_idx, column=2, value=btxn.get("description", "")).border = thin_border
        c = ws_matched.cell(row=row_idx, column=3, value=abs(btxn.get("amount", 0)))
        c.number_format = '$#,##0.00'
        c.border = thin_border
        ws_matched.cell(row=row_idx, column=4, value=inv.get("date", "")).border = thin_border
        ws_matched.cell(row=row_idx, column=5, value=inv.get("vendor", "")).border = thin_border
        c = ws_matched.cell(row=row_idx, column=6, value=inv.get("amount", 0))
        c.number_format = '$#,##0.00'
        c.border = thin_border
        ws_matched.cell(row=row_idx, column=7, value=match.get("match_type", "")).border = thin_border
        c = ws_matched.cell(row=row_idx, column=8, value=match.get("match_score", 0))
        c.number_format = '0.000'
        c.border = thin_border

    for col in range(1, 9):
        ws_matched.column_dimensions[get_column_letter(col)].width = 18

    # ── Tab 3: Unmatched Bank ──
    ws_ubank = wb.create_sheet("Unmatched Bank")
    ws_ubank.sheet_properties.tabColor = "FF9800"

    ws_ubank.merge_cells("A1:F1")
    t = ws_ubank.cell(row=1, column=1, value="Unmatched Bank Transactions (not in email)")
    t.font = Font(bold=True, size=14, color="E65100")
    t.alignment = Alignment(horizontal="center")

    ub_headers = ["Date", "Description", "Amount", "Type", "Category", "Source"]
    for col, h in enumerate(ub_headers, 1):
        cell = ws_ubank.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, txn in enumerate(result.get("unmatched_bank", []), 4):
        ws_ubank.cell(row=row_idx, column=1, value=txn.get("date", "")).border = thin_border
        ws_ubank.cell(row=row_idx, column=2, value=txn.get("description", "")).border = thin_border
        c = ws_ubank.cell(row=row_idx, column=3, value=txn.get("amount", 0))
        c.number_format = '$#,##0.00'
        c.border = thin_border
        ws_ubank.cell(row=row_idx, column=4, value=txn.get("type", "")).border = thin_border
        ws_ubank.cell(row=row_idx, column=5,
                       value=(txn.get("category") or "").replace("_", " ").title()).border = thin_border
        ws_ubank.cell(row=row_idx, column=6, value=txn.get("source_file", "")).border = thin_border

    for col in range(1, 7):
        ws_ubank.column_dimensions[get_column_letter(col)].width = 20

    # ── Tab 4: Unmatched Invoices ──
    ws_uinv = wb.create_sheet("Unmatched Invoices")
    ws_uinv.sheet_properties.tabColor = "9C27B0"

    ws_uinv.merge_cells("A1:E1")
    t = ws_uinv.cell(row=1, column=1, value="Unmatched Invoices (not in bank)")
    t.font = Font(bold=True, size=14, color="6A1B9A")
    t.alignment = Alignment(horizontal="center")

    ui_headers = ["Date", "Vendor", "Amount", "Subject", "Category"]
    for col, h in enumerate(ui_headers, 1):
        cell = ws_uinv.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color="6A1B9A", end_color="6A1B9A", fill_type="solid")
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, inv in enumerate(result.get("unmatched_invoices", []), 4):
        ws_uinv.cell(row=row_idx, column=1, value=inv.get("date", "")).border = thin_border
        ws_uinv.cell(row=row_idx, column=2, value=inv.get("vendor", "")).border = thin_border
        c = ws_uinv.cell(row=row_idx, column=3, value=inv.get("amount", 0))
        c.number_format = '$#,##0.00'
        c.border = thin_border
        ws_uinv.cell(row=row_idx, column=4, value=(inv.get("subject") or "")[:50]).border = thin_border
        ws_uinv.cell(row=row_idx, column=5, value=inv.get("category", "")).border = thin_border

    for col in range(1, 6):
        ws_uinv.column_dimensions[get_column_letter(col)].width = 20

    # ── Tab 5: Cash Flow ──
    ws_cf = wb.create_sheet("Cash Flow")
    ws_cf.sheet_properties.tabColor = "2F5233"

    ws_cf.merge_cells("A1:D1")
    t = ws_cf.cell(row=1, column=1, value="Complete Cash Flow (from Bank)")
    t.font = Font(bold=True, size=16, color="2F5233")
    t.alignment = Alignment(horizontal="center")

    row = 3
    # Income section
    ws_cf.cell(row=row, column=1, value="INCOME").font = Font(bold=True, size=13, color="006600")
    row += 1
    for cat, data in sorted(
        cash_flow.get("income", {}).items(),
        key=lambda x: x[1]["total"], reverse=True,
    ):
        ws_cf.cell(row=row, column=1, value=f"  {cat.replace('_', ' ').title()}")
        ws_cf.cell(row=row, column=2, value=data.get("count", 0))
        c = ws_cf.cell(row=row, column=3, value=data.get("total", 0))
        c.number_format = '$#,##0.00'
        c.font = Font(color="006600")
        row += 1

    total_in = cash_flow.get("total_income", 0)
    ws_cf.cell(row=row, column=1, value="Total Income").font = Font(bold=True, size=12)
    c = ws_cf.cell(row=row, column=3, value=total_in)
    c.font = Font(bold=True, size=12, color="006600")
    c.number_format = '$#,##0.00'
    row += 2

    # Expenses section
    ws_cf.cell(row=row, column=1, value="EXPENSES").font = Font(bold=True, size=13, color="CC0000")
    row += 1
    for cat, data in sorted(
        cash_flow.get("expenses", {}).items(),
        key=lambda x: x[1]["total"],
    ):
        ws_cf.cell(row=row, column=1, value=f"  {cat.replace('_', ' ').title()}")
        ws_cf.cell(row=row, column=2, value=data.get("count", 0))
        c = ws_cf.cell(row=row, column=3, value=abs(data.get("total", 0)))
        c.number_format = '$#,##0.00'
        c.font = Font(color="CC0000")
        row += 1

    total_out = cash_flow.get("total_expenses", 0)
    ws_cf.cell(row=row, column=1, value="Total Expenses").font = Font(bold=True, size=12)
    c = ws_cf.cell(row=row, column=3, value=abs(total_out))
    c.font = Font(bold=True, size=12, color="CC0000")
    c.number_format = '$#,##0.00'
    row += 2

    # Net
    net = cash_flow.get("net_cash_flow", 0)
    ws_cf.cell(row=row, column=1, value="NET CASH FLOW").font = Font(bold=True, size=14)
    np_color = "006600" if net >= 0 else "CC0000"
    c = ws_cf.cell(row=row, column=3, value=net)
    c.font = Font(bold=True, size=14, color=np_color)
    c.number_format = '$#,##0.00'

    for col in range(1, 5):
        ws_cf.column_dimensions[get_column_letter(col)].width = 25

    wb.save(output_path)
    return output_path


# ──────────────────────────────────────────────────────────────
#  Purchase order export
# ──────────────────────────────────────────────────────────────

def export_order_sheet(
    order: dict[str, Any],
    config: Any,
    output_path: Path | None = None,
) -> Path:
    """Export a vendor-ready purchase order to Excel."""
    vendor = order.get("vendor", "Unknown")
    safe_vendor = re.sub(r'[^a-zA-Z0-9]', '_', vendor)

    if output_path is None:
        date_str = datetime.now().strftime("%Y_%m_%d")
        filename = f"PO_{safe_vendor}_{date_str}.xlsx"
        config.ensure_dirs()
        output_path = config.output_dir / filename

    wb = Workbook()
    theme = "2F5233"
    header_font, header_fill, header_alignment, thin_border = _get_styles(theme)

    ws = wb.active
    ws.title = "Purchase Order"
    ws.sheet_properties.tabColor = theme

    # Header
    ws.merge_cells("A1:G1")
    t = ws.cell(row=1, column=1, value="PURCHASE ORDER - Desi Delight")
    t.font = Font(bold=True, size=18, color=theme)
    t.alignment = Alignment(horizontal="center")

    ws.cell(row=3, column=1, value="Vendor:").font = Font(bold=True)
    ws.cell(row=3, column=2, value=vendor).font = Font(bold=True, size=12)
    ws.cell(row=4, column=1, value="Date:").font = Font(bold=True)
    ws.cell(row=4, column=2, value=datetime.now().strftime("%B %d, %Y"))
    ws.cell(row=5, column=1, value="PO #:").font = Font(bold=True)
    ws.cell(row=5, column=2, value=f"PO-{datetime.now().strftime('%Y%m%d')}-{safe_vendor[:8]}")

    # Item table
    po_headers = ["#", "Product", "SKU", "Qty", "Unit", "Unit Cost", "Line Total"]
    for col, h in enumerate(po_headers, 1):
        cell = ws.cell(row=7, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    items = order.get("items", [])
    for i, item in enumerate(items, 1):
        row = 7 + i
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=2, value=item.get("product_name", "")).border = thin_border
        ws.cell(row=row, column=3, value=item.get("sku", "")).border = thin_border
        c = ws.cell(row=row, column=4, value=item.get("quantity", 0))
        c.border = thin_border
        ws.cell(row=row, column=5, value=item.get("unit", "")).border = thin_border
        c = ws.cell(row=row, column=6, value=item.get("unit_cost", 0))
        c.number_format = '$#,##0.00'
        c.border = thin_border
        c = ws.cell(row=row, column=7, value=item.get("line_total", 0))
        c.number_format = '$#,##0.00'
        c.border = thin_border

    # Subtotal row
    total_row = 8 + len(items)
    ws.cell(row=total_row, column=5, value="SUBTOTAL:").font = Font(bold=True, size=12)
    c = ws.cell(row=total_row, column=7, value=order.get("order_total", 0))
    c.font = Font(bold=True, size=12, color=theme)
    c.number_format = '$#,##0.00'

    # Column widths
    widths = [5, 30, 12, 8, 8, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    return output_path


def export_food_cost_report(
    food_cost_data: dict[str, Any],
    trend_data: list[dict[str, Any]],
    config: Any,
    output_path: Path | None = None,
) -> Path:
    """Export food cost analytics to Excel — 2 sheets: Breakdown + Trend."""
    if output_path is None:
        month_str = datetime.now().strftime("%Y_%m")
        filename = f"DD_FOOD_COST_{month_str}.xlsx"
        config.ensure_dirs()
        output_path = config.output_dir / filename

    wb = Workbook()
    theme = "2F5233"
    header_font, header_fill, header_alignment, thin_border = _get_styles(theme)

    # ── Sheet 1: Food Cost ──
    ws = wb.active
    ws.title = "Food Cost"
    ws.sheet_properties.tabColor = theme

    ws.merge_cells("A1:D1")
    t = ws.cell(row=1, column=1, value="Food Cost Analysis")
    t.font = Font(bold=True, size=16, color=theme)
    t.alignment = Alignment(horizontal="center")

    pct = food_cost_data.get("food_cost_pct", 0)
    status = food_cost_data.get("status", "healthy")
    status_color = {"healthy": "006600", "warning": "CC6600", "critical": "CC0000"}.get(status, "000000")

    ws.cell(row=3, column=1, value="Food Cost %:").font = Font(bold=True, size=14)
    ws.cell(row=3, column=2, value=f"{pct}%").font = Font(bold=True, size=14, color=status_color)
    ws.cell(row=4, column=1, value="Status:").font = Font(bold=True)
    ws.cell(row=4, column=2, value=status.upper()).font = Font(color=status_color)
    ws.cell(row=5, column=1, value="Net Sales:").font = Font(bold=True)
    c = ws.cell(row=5, column=2, value=food_cost_data.get("net_sales", 0))
    c.number_format = '$#,##0.00'
    ws.cell(row=6, column=1, value="Food Expenses:").font = Font(bold=True)
    c = ws.cell(row=6, column=2, value=food_cost_data.get("food_cost_total", 0))
    c.number_format = '$#,##0.00'

    cat_headers = ["Category", "Amount", "% of Sales"]
    for col, h in enumerate(cat_headers, 1):
        cell = ws.cell(row=8, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    by_cat = food_cost_data.get("by_category", {})
    row_num = 9
    for cat, info in sorted(by_cat.items(), key=lambda x: x[1].get("total", 0), reverse=True):
        if info.get("total", 0) > 0:
            ws.cell(row=row_num, column=1, value=cat.replace("_", " ").title())
            c = ws.cell(row=row_num, column=2, value=info["total"])
            c.number_format = '$#,##0.00'
            ws.cell(row=row_num, column=3, value=f"{info.get('pct', 0)}%")
            row_num += 1

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # ── Sheet 2: Trend ──
    ws_trend = wb.create_sheet("Trend")
    ws_trend.sheet_properties.tabColor = "1565C0"

    ws_trend.merge_cells("A1:E1")
    t = ws_trend.cell(row=1, column=1, value="Food Cost Trend")
    t.font = Font(bold=True, size=16, color="1565C0")
    t.alignment = Alignment(horizontal="center")

    trend_headers = ["Month", "Net Sales", "Food Cost", "Food Cost %", "Status"]
    for col, h in enumerate(trend_headers, 1):
        cell = ws_trend.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        cell.alignment = header_alignment

    for i, snap in enumerate(trend_data, 4):
        ws_trend.cell(row=i, column=1, value=snap.get("month", ""))
        c = ws_trend.cell(row=i, column=2, value=snap.get("net_sales", 0))
        c.number_format = '$#,##0.00'
        c = ws_trend.cell(row=i, column=3, value=snap.get("food_cost_total", 0))
        c.number_format = '$#,##0.00'
        ws_trend.cell(row=i, column=4, value=f"{snap.get('food_cost_pct', 0)}%")
        ws_trend.cell(row=i, column=5, value=snap.get("status", "").upper())

    for col in range(1, 6):
        ws_trend.column_dimensions[get_column_letter(col)].width = 18

    wb.save(output_path)
    return output_path
